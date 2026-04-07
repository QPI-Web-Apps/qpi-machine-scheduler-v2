"""Stress test initial_tools by manipulating the input Excel directly.

For each scenario, this script:
  1. Copies the latest schedule from ~/Downloads
  2. Mutates specific rows to inject the test condition
     (set IP, mark P+, change tool ID, etc.)
  3. POSTs the modified file to the live API with chosen initial_tools
  4. Inspects the resulting schedule and asserts the expected behavior

Each scenario tests one of the rules from the 2026-04-03 touchbase meeting:
  • In-progress overrides initial_tool entirely
  • Priority chain runs regardless of initial_tool
  • Upfront CO appears before the first JOB on a tool mismatch
  • Self-service vs maintenance changeover types are respected
"""
from __future__ import annotations

import json
import shutil
import sys
import tempfile
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

SRC = Path.home() / "Downloads/04.06 2PMToBeFilled_Updated.xlsx"
BASE = "http://127.0.0.1:8765"


# ── Excel helpers ─────────────────────────────────────────────────────

def _load_headers(path: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    wb.close()
    return {h: i + 1 for i, h in enumerate(headers)}  # 1-indexed for openpyxl


def _find_rows(path: Path, predicate) -> list[tuple[int, dict]]:
    """Return [(row_index, row_dict), ...] for rows matching predicate."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    out = []
    for r_idx in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r_idx]]
        if not row or not row[0]:
            continue
        d = dict(zip(headers, row))
        if predicate(d):
            out.append((r_idx, d))
    wb.close()
    return out


def _make_modified_xlsx(mutations: list[tuple[int, str, object]]) -> Path:
    """Copy SRC and apply (row_index, header_name, new_value) mutations.

    Returns path to the temp file.
    """
    tmp = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    shutil.copy(SRC, tmp)
    wb = openpyxl.load_workbook(tmp)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    h_idx = {h: i + 1 for i, h in enumerate(headers)}
    for row_idx, header, value in mutations:
        col = h_idx[header]
        ws.cell(row=row_idx, column=col, value=value)
    wb.save(tmp)
    wb.close()
    return tmp


# ── API helper ────────────────────────────────────────────────────────

def _post_schedule(xlsx_path: Path, initial_tools: dict) -> dict:
    boundary = "----qpi"
    body = []
    def f(n, v):
        body.append(f"--{boundary}\r\n".encode())
        body.append(f'Content-Disposition: form-data; name="{n}"\r\n\r\n'.encode())
        body.append(f"{v}\r\n".encode())
    f("reference_date", "2026-04-06")
    f("reference_time", "06:30")
    f("max_concurrent", "5")
    f("include_yellow", "true")
    f("include_pink", "true")
    f("shift_config", "{}")
    f("initial_tools", json.dumps(initial_tools))
    body.append(f"--{boundary}\r\n".encode())
    body.append(
        f'Content-Disposition: form-data; name="schedule_file"; filename="{xlsx_path.name}"\r\n'
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n".encode()
    )
    body.append(xlsx_path.read_bytes())
    body.append(b"\r\n")
    body.append(f"--{boundary}--\r\n".encode())
    req = urllib.request.Request(
        f"{BASE}/api/schedule",
        data=b"".join(body),
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=300) as r:
        return json.loads(r.read().decode())


def _machine_entries(d: dict, machine_id: str) -> list[dict]:
    return sorted(
        [e for e in d["schedule"] if e["machine_id"] == machine_id],
        key=lambda e: e["start"],
    )


def _show(d: dict, machine_id: str, n: int = 6) -> None:
    print(f"  {machine_id}:")
    for e in _machine_entries(d, machine_id)[:n]:
        so = f"  so={e.get('so_number','')}" if e["type"] == "JOB" else ""
        print(f"    {e['type']:<11} {e['start'][11:16]} → {e['end'][11:16]}  tool={e.get('tool_id','-')}{so}")


def _check_no_overlap(d: dict, machine_id: str) -> bool:
    """Verify no two non-NOT_RUNNING entries overlap."""
    real = [e for e in _machine_entries(d, machine_id) if e["type"] != "NOT_RUNNING"]
    for i in range(len(real) - 1):
        a_end = datetime.fromisoformat(real[i]["end"])
        b_start = datetime.fromisoformat(real[i + 1]["start"])
        if a_end > b_start + timedelta(seconds=2):
            print(f"  OVERLAP between {real[i]['type']} (ends {real[i]['end']}) and {real[i+1]['type']} (starts {real[i+1]['start']})")
            return False
    return True


# ── Scenarios ────────────────────────────────────────────────────────

def scenario_1_existing_ip_with_init_tool() -> bool:
    """Existing IP on 16B + init_tool[16B] mismatch → no upfront CO (IP overrides)."""
    print("\n" + "=" * 78)
    print("SCENARIO 1: Existing IP on 16B + init_tool[16B]='WRONG' (IP should override)")
    print("=" * 78)
    # The downloaded file already has STF-16ST-B as IP — no mutation needed
    tmp = _make_modified_xlsx([])
    try:
        d = _post_schedule(tmp, {"16B": "WRONG_TOOL"})
    finally:
        tmp.unlink(missing_ok=True)
    _show(d, "16B", 6)
    es = [e for e in _machine_entries(d, "16B") if e["type"] != "NOT_RUNNING"]
    first = es[0] if es else None
    ok = (
        first is not None
        and first["type"] == "JOB"
        and _check_no_overlap(d, "16B")
    )
    print(f"  → {'PASS' if ok else 'FAIL'}: first non-idle is {first['type'] if first else 'none'} (expected JOB; IP overrides init_tool)")
    return ok


def scenario_2_inject_ip_on_16A() -> bool:
    """Force a green job to be IP on 16A. Then init_tool[16A]='WRONG' should be ignored."""
    print("\n" + "=" * 78)
    print("SCENARIO 2: Inject IP on 16A + init_tool[16A]='WRONG' (IP should override)")
    print("=" * 78)
    # Find a Green job assignable to 16A (EQP starting STF-16ST is fine)
    candidates = _find_rows(SRC, lambda d: (
        (d.get("Ticket Color") or "").lower() == "green"
        and "STF-16ST" in (d.get("EQP Code") or "")
        and not (d.get("In Progress") or "")
    ))
    if not candidates:
        print("  SKIP: no candidate row")
        return True
    row_idx, sample = candidates[0]
    print(f"  using row {row_idx}: SO={sample['SO #']}  tool={sample['Tool #']}")
    tmp = _make_modified_xlsx([
        (row_idx, "In Progress", "STF-16ST-A"),  # pin to 16A
    ])
    try:
        d = _post_schedule(tmp, {"16A": "WRONG_TOOL"})
    finally:
        tmp.unlink(missing_ok=True)
    _show(d, "16A", 6)
    es = [e for e in _machine_entries(d, "16A") if e["type"] != "NOT_RUNNING"]
    first = es[0] if es else None
    expected_so = sample["SO #"]
    ok = (
        first is not None
        and first["type"] == "JOB"
        and first.get("so_number") == expected_so
        and _check_no_overlap(d, "16A")
    )
    print(f"  → {'PASS' if ok else 'FAIL'}: first JOB so={first.get('so_number') if first else None}  (expected {expected_so})")
    return ok


def scenario_3_priority_plus_matching_init_tool() -> bool:
    """Mark a job P+ on 16A; set init_tool[16A] to that job's tool. Expect no upfront CO."""
    print("\n" + "=" * 78)
    print("SCENARIO 3: Mark green job P+ on 16A + init_tool=matching tool (no CO expected)")
    print("=" * 78)
    candidates = _find_rows(SRC, lambda d: (
        (d.get("Ticket Color") or "").lower() == "green"
        and "STF-16ST" in (d.get("EQP Code") or "")
        and not (d.get("In Progress") or "")
        and d.get("Tool #")
    ))
    if not candidates:
        print("  SKIP: no candidate")
        return True
    row_idx, sample = candidates[0]
    target_tool = str(sample["Tool #"]).strip()
    print(f"  using row {row_idx}: SO={sample['SO #']} tool={target_tool}; pinning to 16A as P+")
    # Use Picked=STF-16ST-A to lock to 16A, and Priority Status=+ to make it P+
    tmp = _make_modified_xlsx([
        (row_idx, "Picked", "STF-16ST-A"),
        (row_idx, "Priority Status", "+"),
    ])
    try:
        d = _post_schedule(tmp, {"16A": target_tool})
    finally:
        tmp.unlink(missing_ok=True)
    _show(d, "16A", 6)
    es = [e for e in _machine_entries(d, "16A") if e["type"] != "NOT_RUNNING"]
    first = es[0] if es else None
    ok = (
        first is not None
        and first["type"] == "JOB"
        and first.get("so_number") == sample["SO #"]
        and first.get("tool_id") == target_tool
        and _check_no_overlap(d, "16A")
    )
    print(f"  → {'PASS' if ok else 'FAIL'}: first={first['type'] if first else None} so={first.get('so_number') if first else None} tool={first.get('tool_id') if first else None}")
    return ok


def scenario_4_priority_plus_mismatching_init_tool() -> bool:
    """Mark a job P+ on 16A with tool X; set init_tool[16A] to a DIFFERENT tool.
    Expect: upfront CHANGEOVER from init_tool → X, then the P+ job."""
    print("\n" + "=" * 78)
    print("SCENARIO 4: Mark green job P+ on 16A + init_tool=DIFFERENT (CO expected)")
    print("=" * 78)
    candidates = _find_rows(SRC, lambda d: (
        (d.get("Ticket Color") or "").lower() == "green"
        and "STF-16ST" in (d.get("EQP Code") or "")
        and not (d.get("In Progress") or "")
        and d.get("Tool #")
    ))
    if not candidates:
        print("  SKIP: no candidate")
        return True
    row_idx, sample = candidates[0]
    target_tool = str(sample["Tool #"]).strip()
    fake_tool = "ZZ_INITIAL_FAKE"
    print(f"  using row {row_idx}: SO={sample['SO #']} tool={target_tool}; init_tool={fake_tool}")
    tmp = _make_modified_xlsx([
        (row_idx, "Picked", "STF-16ST-A"),
        (row_idx, "Priority Status", "+"),
    ])
    try:
        d = _post_schedule(tmp, {"16A": fake_tool})
    finally:
        tmp.unlink(missing_ok=True)
    _show(d, "16A", 6)
    es = [e for e in _machine_entries(d, "16A") if e["type"] != "NOT_RUNNING"]
    first = es[0] if es else None
    second = es[1] if len(es) > 1 else None
    ok = (
        first is not None
        and first["type"] == "CHANGEOVER"
        and first.get("tool_id") == f"{fake_tool} -> {target_tool}"
        and second is not None
        and second["type"] == "JOB"
        and second.get("so_number") == sample["SO #"]
        and second.get("tool_id") == target_tool
        and _check_no_overlap(d, "16A")
    )
    print(f"  → {'PASS' if ok else 'FAIL'}: first={first['type'] if first else None} tool={first.get('tool_id') if first else None}; second={second['type'] if second else None} so={second.get('so_number') if second else None}")
    return ok


def scenario_5_lmb_tool_swap_with_priority() -> bool:
    """LMB self-service: P+ job + init_tool mismatch should produce TOOL_SWAP, not CHANGEOVER."""
    print("\n" + "=" * 78)
    print("SCENARIO 5: P+ job on LMB + init_tool mismatch (TOOL_SWAP expected)")
    print("=" * 78)
    candidates = _find_rows(SRC, lambda d: (
        (d.get("Ticket Color") or "").lower() == "green"
        and "LMB" in (d.get("EQP Code") or "")
        and d.get("Tool #")
    ))
    if not candidates:
        print("  SKIP: no LMB candidate")
        return True
    row_idx, sample = candidates[0]
    target_tool = str(sample["Tool #"]).strip()
    print(f"  using row {row_idx}: SO={sample['SO #']} tool={target_tool}")
    tmp = _make_modified_xlsx([
        (row_idx, "Priority Status", "+"),
    ])
    try:
        d = _post_schedule(tmp, {"LMB": "ZZ_FAKE_LMB"})
    finally:
        tmp.unlink(missing_ok=True)
    _show(d, "LMB", 6)
    es = [e for e in _machine_entries(d, "LMB") if e["type"] != "NOT_RUNNING"]
    first = es[0] if es else None
    ok = (
        first is not None
        and first["type"] == "TOOL_SWAP"  # not CHANGEOVER
        and first.get("tool_id", "").startswith("ZZ_FAKE_LMB -> ")
        and _check_no_overlap(d, "LMB")
    )
    print(f"  → {'PASS' if ok else 'FAIL'}: first type={first['type'] if first else None} (expected TOOL_SWAP)")
    return ok


def scenario_6_no_jobs_with_matching_tool() -> bool:
    """Set init_tool to a tool that NO 16A job uses. The CO should still appear,
    transitioning from the fake tool to whatever the priority chain picks."""
    print("\n" + "=" * 78)
    print("SCENARIO 6: init_tool[16A]='UNUSED_TOOL' (no 16A job uses it)")
    print("=" * 78)
    tmp = _make_modified_xlsx([])
    try:
        d = _post_schedule(tmp, {"16A": "UNUSED_TOOL_XYZ"})
    finally:
        tmp.unlink(missing_ok=True)
    _show(d, "16A", 6)
    es = [e for e in _machine_entries(d, "16A") if e["type"] != "NOT_RUNNING"]
    first = es[0] if es else None
    second = es[1] if len(es) > 1 else None
    ok = (
        first is not None
        and first["type"] == "CHANGEOVER"
        and first.get("tool_id", "").startswith("UNUSED_TOOL_XYZ -> ")
        and second is not None
        and second["type"] == "JOB"
        and _check_no_overlap(d, "16A")
    )
    print(f"  → {'PASS' if ok else 'FAIL'}: first={first['type'] if first else None} tool={first.get('tool_id') if first else None}")
    return ok


def scenario_7_machine_20_no_changeover() -> bool:
    """Machine 20 has no changeovers — init_tool should be ignored."""
    print("\n" + "=" * 78)
    print("SCENARIO 7: Machine 20 (no CO) + init_tool='WHATEVER' (should be ignored)")
    print("=" * 78)
    tmp = _make_modified_xlsx([])
    try:
        d = _post_schedule(tmp, {"20": "WHATEVER_TOOL"})
    finally:
        tmp.unlink(missing_ok=True)
    _show(d, "20", 6)
    es = [e for e in _machine_entries(d, "20") if e["type"] != "NOT_RUNNING"]
    first = es[0] if es else None
    has_any_co = any(e["type"] in ("CHANGEOVER", "TOOL_SWAP") for e in es)
    ok = first and first["type"] == "JOB" and not has_any_co
    print(f"  → {'PASS' if ok else 'FAIL'}: first={first['type'] if first else None}, any CO entry={has_any_co}")
    return ok


def scenario_8_p_plus_on_one_tool_init_on_other() -> bool:
    """Inject TWO P+ jobs on 16A with different tools, init_tool=neither.
    Expect: solver picks one of them as first, CO inserted with the chosen tool."""
    print("\n" + "=" * 78)
    print("SCENARIO 8: Two P+ jobs on 16A with different tools + init_tool=other")
    print("=" * 78)
    candidates = _find_rows(SRC, lambda d: (
        (d.get("Ticket Color") or "").lower() == "green"
        and "STF-16ST" in (d.get("EQP Code") or "")
        and not (d.get("In Progress") or "")
        and d.get("Tool #")
    ))
    if len(candidates) < 2:
        print("  SKIP: need 2 candidates")
        return True
    # Pick two with DIFFERENT tools
    seen_tools = set()
    picks = []
    for r, d in candidates:
        t = str(d["Tool #"]).strip()
        if t not in seen_tools:
            picks.append((r, d, t))
            seen_tools.add(t)
        if len(picks) == 2:
            break
    if len(picks) < 2:
        print("  SKIP: couldn't find 2 jobs with distinct tools")
        return True
    (r1, s1, t1), (r2, s2, t2) = picks
    print(f"  P+ #1: row {r1}  SO={s1['SO #']} tool={t1}")
    print(f"  P+ #2: row {r2}  SO={s2['SO #']} tool={t2}")
    tmp = _make_modified_xlsx([
        (r1, "Picked", "STF-16ST-A"),
        (r1, "Priority Status", "+"),
        (r2, "Picked", "STF-16ST-A"),
        (r2, "Priority Status", "+"),
    ])
    try:
        d = _post_schedule(tmp, {"16A": "ZZ_INIT"})
    finally:
        tmp.unlink(missing_ok=True)
    _show(d, "16A", 8)
    es = [e for e in _machine_entries(d, "16A") if e["type"] != "NOT_RUNNING"]
    first = es[0] if es else None
    # Whichever job runs first, the CO should reference its tool
    expected_tools = {t1, t2}
    first_so_set = {s1["SO #"], s2["SO #"]}
    first_job = next((e for e in es if e["type"] == "JOB"), None)
    ok = (
        first is not None
        and first["type"] == "CHANGEOVER"
        and first_job is not None
        and first_job.get("tool_id") in expected_tools
        and first_job.get("so_number") in first_so_set
        and first.get("tool_id") == f"ZZ_INIT -> {first_job.get('tool_id')}"
        and _check_no_overlap(d, "16A")
    )
    print(f"  → {'PASS' if ok else 'FAIL'}: first={first['type'] if first else None}; first JOB so={first_job.get('so_number') if first_job else None} tool={first_job.get('tool_id') if first_job else None}")
    return ok


def main() -> int:
    if not SRC.exists():
        print(f"FAIL: source xlsx not found at {SRC}")
        return 1

    scenarios = [
        scenario_1_existing_ip_with_init_tool,
        scenario_2_inject_ip_on_16A,
        scenario_3_priority_plus_matching_init_tool,
        scenario_4_priority_plus_mismatching_init_tool,
        scenario_5_lmb_tool_swap_with_priority,
        scenario_6_no_jobs_with_matching_tool,
        scenario_7_machine_20_no_changeover,
        scenario_8_p_plus_on_one_tool_init_on_other,
    ]

    results = []
    for fn in scenarios:
        try:
            results.append((fn.__name__, fn()))
        except Exception as e:
            print(f"  EXCEPTION: {type(e).__name__}: {e}")
            results.append((fn.__name__, False))

    print("\n" + "=" * 78)
    print("SUMMARY")
    print("=" * 78)
    for name, ok in results:
        mark = "✓" if ok else "✗"
        print(f"  {mark} {name}")
    n_pass = sum(1 for _, ok in results if ok)
    print(f"\n{n_pass}/{len(results)} scenarios passed")
    return 0 if n_pass == len(results) else 1


if __name__ == "__main__":
    sys.exit(main())
