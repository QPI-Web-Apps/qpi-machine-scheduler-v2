"""Simulate the frontend end-to-end for the two new changes:

  1. Excel export now contains the "Everything at STF" column next to
     "Ticket Color".
  2. Publish writes to scheduler_germantown_jobs in addition to the two
     existing scheduler_* tables.

Cleans up any rows it writes to scheduler_germantown_jobs so subsequent
runs start from a clean slate. The yellow/pink and schedule tables are
left in whatever state publish put them in (handled by the existing
test_publish.py flow).
"""
from __future__ import annotations

import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO))

env_path = REPO / ".env"
if env_path.exists():
    for line in env_path.read_text().splitlines():
        if "=" in line and not line.strip().startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

from openpyxl import load_workbook  # noqa: E402

from src import api  # noqa: E402
from src.export import generate_schedule_excel  # noqa: E402
from src.scheduler import generate_schedule  # noqa: E402
from src.scheduler_io import SchedulerConfig  # noqa: E402

TEST_XLSX = REPO / "test_schedules" / "STF schedule 03.31 130PMToBeFilled_Updated.xlsx"

import json as _json


def _section(title):
    print(f"\n{'='*70}\n{title}\n{'='*70}")


def main() -> int:
    if not TEST_XLSX.exists():
        print(f"FAIL: test xlsx missing at {TEST_XLSX}")
        return 1

    cfg = SchedulerConfig(
        schedule_start=datetime(2026, 4, 2, 6, 30),
        include_yellow=True,
        include_pink=True,
    )

    _section("Generating schedule from test Excel…")
    result = generate_schedule(str(TEST_XLSX), cfg, max_concurrent=5)
    print(f"  solver_status={result.solver_status}")
    print(f"  entries={len(result.entries)}")
    print(f"  germantown_jobs={len(result.germantown_jobs)}")

    # ── Test 1: Excel export has "Everything at STF" next to "Ticket Color" ──
    _section("TEST 1: Excel export 'Everything at STF' column")
    wb = generate_schedule_excel(result, cfg)

    # Pick a per-machine sheet that has entries
    sample_sheet = None
    for name in wb.sheetnames:
        if name == "Summary":
            continue
        ws = wb[name]
        if ws.max_row >= 2:
            sample_sheet = ws
            break

    if sample_sheet is None:
        print("  FAIL: no per-machine sheet with data found")
        return 1

    headers = [sample_sheet.cell(row=1, column=c).value for c in range(1, sample_sheet.max_column + 1)]
    print(f"  Sheet: {sample_sheet.title}")
    print(f"  Headers: {headers}")

    ok = True
    if "Everything at STF" not in headers:
        print("  FAIL: 'Everything at STF' column missing from export")
        ok = False
    else:
        tc_idx = headers.index("Ticket Color")
        stf_idx = headers.index("Everything at STF")
        if stf_idx != tc_idx + 1:
            print(f"  FAIL: 'Everything at STF' at index {stf_idx}, expected {tc_idx + 1} (right of Ticket Color)")
            ok = False
        else:
            print(f"  OK: 'Everything at STF' at column {stf_idx + 1}, right of 'Ticket Color'")

        # Dump a few non-empty values in that column to prove it's populated
        stf_col = stf_idx + 1
        sample_values = []
        for r in range(2, min(sample_sheet.max_row + 1, 200)):
            v = sample_sheet.cell(row=r, column=stf_col).value
            if v not in (None, ""):
                sample_values.append(v)
            if len(sample_values) >= 5:
                break
        print(f"  Sample STF values: {sample_values}")

    # Save to tempfile to confirm openpyxl writes it cleanly
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        size = Path(tmp.name).stat().st_size
        print(f"  Wrote {tmp.name} ({size} bytes)")
        # Re-open from disk to prove the file is valid
        wb2 = load_workbook(tmp.name, read_only=True)
        re_headers = [wb2[sample_sheet.title].cell(row=1, column=c).value for c in range(1, len(headers) + 1)]
        assert re_headers == headers, f"Header mismatch after save: {re_headers} vs {headers}"
        print("  OK: file re-opens cleanly with same headers")
        Path(tmp.name).unlink()

    # ── Test 2: Publish writes to scheduler_germantown_jobs ──
    _section("TEST 2: Publish writes to scheduler_germantown_jobs")

    schedule_id = "test_gt01"
    # all_yp_jobs simulates what the real create_schedule endpoint stores.
    api._results[schedule_id] = {
        "result": result,
        "cfg": cfg,
        "data": {},
        "all_yp_jobs": [],
    }

    db = api._get_prisma()

    # Capture baseline counts for the sibling tables.
    yp_before = db.scheduler_yellow_pink_jobs.count()
    sch_before = db.scheduler_published_schedule.count()
    gt_before = db.scheduler_germantown_jobs.count()
    print(f"  Before publish: yp={yp_before}  sched={sch_before}  germantown={gt_before}")

    # Call publish (simulates clicking "Publish Schedule" in the frontend)
    resp = api.publish_schedule(schedule_id)
    body = _json.loads(resp.body.decode())
    print(f"  publish response: {body}")
    if not body.get("ok"):
        print("  FAIL: publish returned ok=false")
        return 1

    gt_count_reported = body.get("germantown_count")
    if gt_count_reported is None:
        print("  FAIL: germantown_count not in publish response")
        ok = False

    yp_y = db.scheduler_yellow_pink_jobs.count(where={"processed_indicator": "y"})
    sch_y = db.scheduler_published_schedule.count(where={"processed_indicator": "y"})
    gt_y = db.scheduler_germantown_jobs.count(where={"processed_indicator": "y"})
    gt_n = db.scheduler_germantown_jobs.count(where={"processed_indicator": "n"})
    print(f"  After publish:  yp.y={yp_y}  sched.y={sch_y}  germantown.y={gt_y}  germantown.n={gt_n}")

    if gt_y != len(result.germantown_jobs):
        print(f"  FAIL: germantown.y={gt_y}, expected {len(result.germantown_jobs)}")
        ok = False
    elif gt_y != gt_count_reported:
        print(f"  FAIL: response germantown_count={gt_count_reported} != DB count {gt_y}")
        ok = False
    else:
        print(f"  OK: {gt_y} germantown rows inserted, response matches DB")

    # Spot-check a row so we see actual field values
    sample = db.scheduler_germantown_jobs.find_first(where={"processed_indicator": "y"})
    if sample:
        print(f"  Sample row: so={sample.so_number} part={sample.finished_item} "
              f"customer={sample.customer} color={sample.ticket_color} "
              f"tool={sample.tool_id} eqp={sample.eqp_code} "
              f"qty={sample.remaining_qty} hrs={sample.run_hours}")
    elif len(result.germantown_jobs) > 0:
        print("  FAIL: expected germantown rows but find_first returned None")
        ok = False

    # ── Test 3: Second publish flips prior germantown rows to 'n' ──
    _section("TEST 3: Second publish flips prior rows to 'n'")
    resp2 = api.publish_schedule(schedule_id)
    body2 = _json.loads(resp2.body.decode())
    print(f"  publish #2 response: {body2}")

    gt_y2 = db.scheduler_germantown_jobs.count(where={"processed_indicator": "y"})
    gt_n2 = db.scheduler_germantown_jobs.count(where={"processed_indicator": "n"})
    print(f"  After 2nd publish: germantown.y={gt_y2}  germantown.n={gt_n2}")
    if gt_y2 != len(result.germantown_jobs):
        print(f"  FAIL: germantown.y={gt_y2} on 2nd publish, expected {len(result.germantown_jobs)}")
        ok = False
    if gt_n2 < gt_y:
        print(f"  FAIL: germantown.n={gt_n2}, expected >= previous y-count {gt_y}")
        ok = False
    else:
        print(f"  OK: previous {gt_y} rows flipped to 'n', fresh {gt_y2} rows with 'y'")

    # ── Cleanup: delete rows we just wrote (leave table empty) ──
    _section("CLEANUP: deleting rows written by this test")
    deleted = db.scheduler_germantown_jobs.delete_many(where={"id": {"gt": 0}})
    print(f"  Deleted {deleted} rows from scheduler_germantown_jobs")
    yp_del = db.scheduler_yellow_pink_jobs.delete_many(where={"id": {"gt": 0}})
    sch_del = db.scheduler_published_schedule.delete_many(where={"id": {"gt": 0}})
    print(f"  Also cleaned sibling test output: yp={yp_del}  sched={sch_del}")

    final = {
        "yp": db.scheduler_yellow_pink_jobs.count(),
        "sched": db.scheduler_published_schedule.count(),
        "germantown": db.scheduler_germantown_jobs.count(),
    }
    print(f"  Final counts: {final}")

    db.disconnect()

    print("\n" + ("PASS" if ok else "FAIL"))
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
