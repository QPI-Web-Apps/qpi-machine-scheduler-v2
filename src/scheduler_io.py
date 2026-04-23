"""Excel loading, job preparation, and scheduler configuration."""

from __future__ import annotations

import math
from dataclasses import dataclass, field
from datetime import datetime, time
from typing import Optional

import openpyxl

from .helpers import (
    DEFAULT_HEADCOUNT,
    PriorityClass,
    classify_priority,
    compute_run_hours,
    infer_machine_from_eqp,
    infer_station_group,
    normalize_tool,
    parse_boolish,
    parse_due_date,
)
from .models import MACHINE_BY_ID, machines_in_group


# ── Config ──────────────────────────────────────────────────────────

@dataclass
class SchedulerConfig:
    schedule_start: datetime
    shifts_per_day: dict[str, int] = field(default_factory=lambda: {
        "16A": 2, "16B": 2, "16C": 2, "20": 2, "8": 2,
        "LMB": 2, "SMB": 2, "6ST": 2, "RF": 2,
    })
    # Per-machine-per-day shift selection: machine_id -> date_str -> [shift_numbers]
    # Empty dict = use shifts_per_day fallback
    shift_schedule: dict[str, dict[str, list[int]]] = field(default_factory=dict)
    include_yellow: bool = False
    include_pink: bool = False
    include_white: bool = False
    include_germantown: bool = False
    disabled_stations: list[str] = field(default_factory=list)
    disabled_machines: list[str] = field(default_factory=list)
    default_headcount: float = DEFAULT_HEADCOUNT
    h3_enabled: bool = True
    initial_tools: dict[str, str] = field(default_factory=dict)
    priority_boost: bool = True
    minimize_late: bool = False
    hc_penalty_weight: float = 30  # minutes-equivalent cost per HC unit of change
    total_crew: int = 0  # max total headcount across concurrent jobs (0 = unlimited, hard constraint)

    def get_day_shift_map(self, machine_id: str):
        """Return per-day shift map for a machine, or int fallback."""
        if machine_id in self.shift_schedule and self.shift_schedule[machine_id]:
            return self.shift_schedule[machine_id]
        return self.shifts_per_day.get(machine_id, 2)


# ── Job dict structure ──────────────────────────────────────────────
# Each job is a plain dict with these keys:
#   so_number, sol_id, finished_item, description, customer,
#   total_qty, produced_qty, remaining_qty, eqp_code,
#   station_group, preferred_machine, tool_id, run_hours,
#   headcount, headcount_assumed, due_date, priority_class,
#   is_labeler, is_bagger, ticket_color, priority_str,
#   is_picked, is_in_progress, order_entry_date,
#   eligible_machines (list[str])


# ── Column mapping ──────────────────────────────────────────────────

_COL_MAP = {
    "SO #": "so_number",
    "SOL ID": "sol_id",
    "Finished Item": "finished_item",
    "Description": "description",
    "Customer": "customer",
    "Total QTY": "total_qty",
    "Produced QTY": "produced_qty",
    "Remaining QTY": "remaining_qty",
    "EQP Code": "eqp_code",
    "Due Date": "due_date",
    "Tool #": "tool_id",
    "Ticket Color": "ticket_color",
    "Priority Status": "priority_str",
    "avg_num_employees": "avg_num_employees",
    "person hour rate": "person_hour_rate",
    "order entry date": "order_entry_date",
    "In progress": "is_in_progress",
    "In Progress": "is_in_progress",
    "in progress": "is_in_progress",
    "Picked Jobs": "is_picked",
    "Picked": "is_picked",
    "picked": "is_picked",
    "label_indicator": "is_labeler",
    "bag_indicator": "is_bagger",
    "Part Number": "part_number",
    "Everything at STF": "at_stf",
}


def _safe_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    return str(val).strip()


def _safe_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        try:
            return float(val)
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        if math.isnan(val):
            return None
        return float(val)
    return None


# ── Loading ─────────────────────────────────────────────────────────

def load_jobs_from_excel(path: str, cfg: SchedulerConfig) -> tuple[list[dict], list[dict], list[dict]]:
    """Load and prepare jobs from an Excel file.

    Returns (jobs, skipped, germantown_jobs) where each is a list of dicts.
    Germantown jobs are green-ticket jobs with "Everything at STF" = N.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in _COL_MAP:
            col_idx[_COL_MAP[h]] = i

    jobs: list[dict] = []
    skipped: list[dict] = []
    germantown_jobs: list[dict] = []

    for row in rows[1:]:
        raw = {key: row[idx] if idx < len(row) else None for key, idx in col_idx.items()}

        so_number = _safe_str(raw.get("so_number"))
        eqp_code = _safe_str(raw.get("eqp_code"))
        if not so_number or not eqp_code:
            skipped.append({"reason": "missing SO# or EQP", "so_number": so_number, "eqp_code": eqp_code})
            continue

        # Station group
        station_group = infer_station_group(eqp_code)
        if not station_group:
            skipped.append({"reason": f"unknown EQP pattern: {eqp_code}", "so_number": so_number})
            continue

        # Disabled stations
        if station_group in cfg.disabled_stations:
            skipped.append({"reason": f"station {station_group} disabled", "so_number": so_number})
            continue

        # Ticket color filtering — only green by default
        color = _safe_str(raw.get("ticket_color")).lower()
        if "green" not in color:
            if "pink" in color and not cfg.include_pink:
                skipped.append({"reason": "pink ticket excluded", "so_number": so_number})
                continue
            if "white" in color and not cfg.include_white:
                skipped.append({"reason": "white ticket excluded", "so_number": so_number})
                continue
            if "yellow" in color and not cfg.include_yellow:  # noqa: SIM114
                skipped.append({"reason": "yellow ticket excluded", "so_number": so_number})
                continue
            if not color:
                skipped.append({"reason": "no ticket color", "so_number": so_number})
                continue

        # Tool
        raw_tool = _safe_str(raw.get("tool_id"))
        if not raw_tool:
            if station_group == "rf":
                raw_tool = "99999"
            else:
                skipped.append({"reason": "blank tool", "so_number": so_number})
                continue
        try:
            tool_id = normalize_tool(raw_tool)
        except ValueError:
            skipped.append({"reason": f"bad tool: {raw_tool}", "so_number": so_number})
            continue

        # Run hours
        remaining_qty = _safe_float(raw.get("remaining_qty")) or 0.0
        person_hour_rate = _safe_float(raw.get("person_hour_rate")) or 0.0
        avg_hc = _safe_float(raw.get("avg_num_employees"))

        run_hours, headcount, hc_assumed = compute_run_hours(
            remaining_qty, person_hour_rate, avg_hc, cfg.default_headcount
        )
        if run_hours <= 0:
            skipped.append({"reason": "zero run hours", "so_number": so_number})
            continue

        # Due date & priority
        due_date = parse_due_date(raw.get("due_date"))
        priority_str = _safe_str(raw.get("priority_str"))
        priority_class = classify_priority(priority_str, due_date, cfg.schedule_start)

        # Flags
        is_labeler = parse_boolish(raw.get("is_labeler"))
        is_bagger = parse_boolish(raw.get("is_bagger"))

        # Parse In-progress / Picked as EQP codes → machine locks
        raw_ip = _safe_str(raw.get("is_in_progress"))
        raw_pk = _safe_str(raw.get("is_picked"))
        in_progress_machine = infer_machine_from_eqp(raw_ip) if raw_ip else None
        picked_machine = infer_machine_from_eqp(raw_pk) if raw_pk else None
        is_in_progress = in_progress_machine is not None
        is_picked = picked_machine is not None and not is_in_progress
        locked_machine = in_progress_machine or picked_machine

        # Override priority for locked jobs
        if is_in_progress:
            priority_class = int(PriorityClass.IN_PROGRESS)
        elif is_picked:
            priority_class = int(PriorityClass.PRIORITY_PLUS)

        # Machine eligibility
        preferred_machine = infer_machine_from_eqp(eqp_code)
        eligible = _build_eligibility(station_group, is_labeler, preferred_machine)
        if not eligible:
            skipped.append({"reason": "no eligible machines", "so_number": so_number})
            continue

        # Lock to specific machine if picked/in-progress
        if locked_machine and locked_machine in eligible:
            eligible = [locked_machine]

        # Remove disabled machines
        if cfg.disabled_machines:
            eligible = [m for m in eligible if m not in cfg.disabled_machines]
            if not eligible:
                skipped.append({"reason": "all eligible machines disabled", "so_number": so_number})
                continue

        # STF check for germantown filtering
        at_stf = _safe_str(raw.get("at_stf")).upper()

        job = {
            "so_number": so_number,
            "sol_id": raw.get("sol_id"),
            "finished_item": _safe_str(raw.get("finished_item")),
            "description": _safe_str(raw.get("description")),
            "customer": _safe_str(raw.get("customer")),
            "total_qty": _safe_float(raw.get("total_qty")) or 0,
            "produced_qty": _safe_float(raw.get("produced_qty")) or 0,
            "remaining_qty": remaining_qty,
            "eqp_code": eqp_code,
            "station_group": station_group,
            "preferred_machine": preferred_machine,
            "tool_id": tool_id,
            "run_hours": run_hours,
            "headcount": headcount,
            "headcount_assumed": hc_assumed,
            "due_date": due_date,
            "priority_class": int(priority_class),
            "is_labeler": is_labeler,
            "is_bagger": is_bagger,
            "ticket_color": _safe_str(raw.get("ticket_color")),
            "priority_str": priority_str,
            "is_picked": is_picked,
            "is_in_progress": is_in_progress,
            "locked_machine": locked_machine,
            "order_entry_date": raw.get("order_entry_date"),
            "eligible_machines": eligible,
            "at_stf": at_stf,
        }

        # Germantown filter: green jobs with STF=N excluded by default
        if at_stf == "N" and "green" in color and not cfg.include_germantown:
            germantown_jobs.append(job)
            continue

        jobs.append(job)

    return jobs, skipped, germantown_jobs


def _build_eligibility(
    station_group: str,
    is_labeler: bool,
    preferred_machine: Optional[str],
) -> list[str]:
    """Determine which machines can run this job."""
    group_machines = machines_in_group(station_group)
    if not group_machines:
        return []

    # Labeler jobs within 16-group → only 16C
    if is_labeler and station_group == "16":
        return ["16C"]

    return [m.machine_id for m in group_machines]
