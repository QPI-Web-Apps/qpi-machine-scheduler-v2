"""Excel export — generate .xlsx workbook from a ScheduleResult."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from .models import MACHINES, MACHINE_BY_ID
from .scheduler import ScheduleResult, compute_machine_summary
from .scheduler_io import SchedulerConfig

# ── Column definitions ─────────────────────────────────────────────

COLUMNS = [
    ("Seq", 6),
    ("Type", 14),
    ("Start", 18),
    ("End", 18),
    ("Shift", 6),
    ("SO #", 12),
    ("Tool ID", 10),
    ("Finished Item", 20),
    ("Description", 30),
    ("Customer", 22),
    ("Remaining QTY", 14),
    ("Run Hrs", 9),
    ("Headcount", 10),
    ("Due Date", 14),
    ("Priority", 8),
    ("Ticket Color", 12),
    ("Everything at STF", 18),
    ("Labeler", 8),
    ("Bag Sealer", 10),
    ("Idle Type", 14),
    ("Crew From", 10),
    ("Crew To", 10),
    ("Late", 6),
    ("Days Early/Late", 14),
]

COL_NAMES = [c[0] for c in COLUMNS]
COL_WIDTHS = [c[1] for c in COLUMNS]

SUMMARY_COLUMNS = [
    ("Machine", 12),
    ("Jobs", 8),
    ("Job Hours", 10),
    ("CO Hours", 10),
    ("Idle Hours", 10),
    ("No Crew Hours", 14),
    ("Total Hours", 12),
    ("Utilization %", 12),
    ("Start", 18),
    ("End", 18),
]

# ── Styles ─────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

DAY_SEP_FILL = PatternFill("solid", fgColor="DBEAFE")
DAY_SEP_FONT = Font(bold=True, size=10)

TYPE_FILLS = {
    "JOB": PatternFill("solid", fgColor="DCFCE7"),
    "CHANGEOVER": PatternFill("solid", fgColor="FED7AA"),
    "TOOL_SWAP": PatternFill("solid", fgColor="FEF9C3"),
    "NOT_RUNNING": PatternFill("solid", fgColor="F1F5F9"),
}

LATE_FONT = Font(color="DC2626", bold=True)

PRIORITY_LABELS = {-1: "In Prog", 0: "P+", 1: "P", 2: "Past Due", 3: "Normal"}


# ── Public API ─────────────────────────────────────────────────────

def generate_schedule_excel(
    result: ScheduleResult,
    cfg: SchedulerConfig,
) -> Workbook:
    """Build a multi-sheet Excel workbook from a schedule result."""
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    # Group entries by machine
    by_machine: dict[str, list] = defaultdict(list)
    for e in result.entries:
        by_machine[e.machine_id].append(e)

    # Summary sheet first
    _write_summary_sheet(wb, result, cfg, by_machine)

    # Per-machine sheets
    for spec in MACHINES:
        entries = sorted(by_machine.get(spec.machine_id, []), key=lambda e: e.start)
        _write_machine_sheet(wb, spec.display_name, entries)

    return wb


# ── Summary sheet ──────────────────────────────────────────────────

def _write_summary_sheet(wb, result, cfg, by_machine):
    ws = wb.create_sheet("Summary")

    # Header
    for col_idx, (name, width) in enumerate(SUMMARY_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    totals = {"jobs": 0, "job_hrs": 0.0, "co_hrs": 0.0,
              "no_crew_hrs": 0.0, "total_hrs": 0.0}

    row = 2
    for spec in MACHINES:
        s = compute_machine_summary(result.entries, spec.machine_id, cfg)

        start_str = s.start.strftime("%m/%d/%Y %H:%M") if s.start else ""
        end_str = s.end.strftime("%m/%d/%Y %H:%M") if s.end else ""

        values = [
            spec.display_name, s.jobs,
            s.job_hours, s.changeover_hours, 0,  # idle_hours placeholder
            s.no_crew_hours, s.total_hours, s.utilization,
            start_str, end_str,
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row, column=col_idx, value=val)

        totals["jobs"] += s.jobs
        totals["job_hrs"] += s.job_hours
        totals["co_hrs"] += s.changeover_hours
        totals["no_crew_hrs"] += s.no_crew_hours
        totals["total_hrs"] += s.total_hours
        row += 1

    # Total row
    total_util = round(totals["job_hrs"] / totals["total_hrs"] * 100, 1) if totals["total_hrs"] > 0 else 0
    total_values = [
        "TOTAL", totals["jobs"],
        round(totals["job_hrs"], 1), round(totals["co_hrs"], 1),
        0, round(totals["no_crew_hrs"], 1),
        round(totals["total_hrs"], 1), total_util, "", "",
    ]
    for col_idx, val in enumerate(total_values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(bold=True)


# ── Per-machine sheet ──────────────────────────────────────────────

def _write_machine_sheet(wb, sheet_name, entries):
    ws = wb.create_sheet(sheet_name)

    # Header row
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    if not entries:
        ws.cell(row=2, column=1, value="No entries scheduled.")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
        ws.cell(row=2, column=1).font = Font(italic=True, color="64748B")
        return

    row = 2
    prev_date = None
    seq = 0

    for entry in entries:
        entry_date = entry.start.date()

        # Day separator
        if prev_date is not None and entry_date != prev_date:
            label = entry_date.strftime("%A, %B %-d, %Y")
            cell = ws.cell(row=row, column=1, value=label)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
            cell.fill = DAY_SEP_FILL
            cell.font = DAY_SEP_FONT
            cell.alignment = Alignment(horizontal="center")
            row += 1
        elif prev_date is None:
            # First day header
            label = entry_date.strftime("%A, %B %-d, %Y")
            cell = ws.cell(row=row, column=1, value=label)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
            cell.fill = DAY_SEP_FILL
            cell.font = DAY_SEP_FONT
            cell.alignment = Alignment(horizontal="center")
            row += 1

        prev_date = entry_date
        seq += 1
        jd = entry.job_data or {}

        # Compute late / days early-late
        late = ""
        days_delta = ""
        if entry.entry_type == "JOB" and jd.get("due_date"):
            due = jd["due_date"]
            if isinstance(due, datetime):
                due_date = due.date()
            else:
                due_date = due
            end_date = entry.end.date()
            delta = (due_date - end_date).days
            late = "Yes" if delta < 0 else "No"
            days_delta = delta

        priority_raw = jd.get("priority_class", "")
        priority_label = PRIORITY_LABELS.get(priority_raw, priority_raw) if priority_raw != "" else ""

        values = [
            seq,
            entry.entry_type,
            entry.start.strftime("%m/%d %H:%M"),
            entry.end.strftime("%m/%d %H:%M"),
            entry.shift,
            entry.so_number or "",
            entry.tool_id or "",
            jd.get("finished_item", ""),
            jd.get("description", ""),
            jd.get("customer", ""),
            jd.get("remaining_qty", ""),
            jd.get("run_hours", ""),
            entry.headcount if entry.headcount is not None else "",
            jd["due_date"].strftime("%m/%d/%Y") if jd.get("due_date") and isinstance(jd["due_date"], datetime) else (str(jd["due_date"]) if jd.get("due_date") else ""),
            priority_label,
            jd.get("ticket_color", ""),
            jd.get("at_stf", ""),
            "Yes" if jd.get("is_labeler") else "",
            "Yes" if jd.get("is_bagger") else "",
            entry.idle_type or "",
            entry.crew_from or "",
            entry.crew_to or "",
            late,
            days_delta,
        ]

        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row, column=col_idx, value=val)

        # Row fill by entry type
        fill = TYPE_FILLS.get(entry.entry_type)
        if fill:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row, column=col_idx).fill = fill

        # Late highlighting
        if late == "Yes":
            for col_idx in (COL_NAMES.index("Late") + 1, COL_NAMES.index("Days Early/Late") + 1):
                ws.cell(row=row, column=col_idx).font = LATE_FONT

        row += 1
